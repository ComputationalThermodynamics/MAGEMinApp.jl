#!/usr/bin/env python3
"""
Generate per-sample CSV files with apfu compositions for Pl_Ca, Amp_Mg, Aug_Mg, Ep_Mg.

Data sources:
  Pl_Ca  : Gen-Oxide-norm-Pl.xls  / 'General oxygen' sheet / Atoms section / Ca row
  Amp_Mg : Amphibole-norm1-serie1.xls + _Toulouse.xlsx / 'NAMP' sheet / Cations all Fe2+ / Mg row
  Aug_Mg : Pyroxene-norm.xls / 'WB-cpx' sheet / Cations all Fe2+ / Mg row
  Ep_Mg  : BenedicteCenki…calc.xlsx / Feuil2 EP section / calculated from oxides (12.5 O)

Output: one CSV per sample in sample_names.txt
  Header row + mean row + std-dev row  (NaN where mineral absent)
"""
import os
import re
import openpyxl
import xlrd
import numpy as np

WORKDIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(WORKDIR)

# ── Oxide normalization constants (from Excel headers) ──────────────────────
OXIDE_MW = {
    'SiO2': 60.08, 'TiO2': 79.9, 'Cr2O3': 151.9908, 'Al2O3': 101.94,
    'FeO': 71.85, 'Fe2O3': 159.691, 'MnO': 70.937, 'NiO': 74.7094,
    'MgO': 40.31, 'CaO': 56.08, 'Na2O': 61.982, 'K2O': 94.2,
    'H2O': 18.01534, 'F': 18.9984, 'Cl': 35.453,
}
OXIDE_CAT = {
    'SiO2': 1, 'TiO2': 1, 'Cr2O3': 2, 'Al2O3': 2, 'FeO': 1, 'Fe2O3': 2,
    'MnO': 1, 'NiO': 1, 'MgO': 1, 'CaO': 1, 'Na2O': 2, 'K2O': 2,
    'H2O': 2, 'F': 1, 'Cl': 1,
}
OXIDE_CHARGE = {
    'SiO2': 4, 'TiO2': 4, 'Cr2O3': 3, 'Al2O3': 3, 'FeO': 2, 'Fe2O3': 3,
    'MnO': 2, 'NiO': 2, 'MgO': 2, 'CaO': 2, 'Na2O': 1, 'K2O': 1,
    'H2O': 0, 'F': 0, 'Cl': 0,
}


def oxygens_per_mol(oxide):
    """Number of oxygens per mole of oxide formula unit."""
    return (OXIDE_CHARGE.get(oxide, 0) * OXIDE_CAT.get(oxide, 1)) / 2


def calc_apfu_from_oxides(oxide_wt, n_oxygens, element):
    """
    Calculate element apfu by oxygen normalization.
    oxide_wt  : dict {oxide_name: wt_percent}
    n_oxygens : target oxygen count (e.g. 12.5 for epidote)
    element   : 'Mg' or 'Ca'
    """
    total_O = 0.0
    for ox, w in oxide_wt.items():
        if ox in OXIDE_MW:
            try:
                fw = float(w)
                if fw > 0:
                    total_O += fw / OXIDE_MW[ox] * oxygens_per_mol(ox)
            except (ValueError, TypeError):
                pass
    if total_O <= 0:
        return np.nan
    scale = n_oxygens / total_O
    ox_map = {'Mg': 'MgO', 'Ca': 'CaO', 'Si': 'SiO2'}
    oxide = ox_map.get(element)
    if oxide:
        try:
            w = float(oxide_wt.get(oxide, 0))
            if w > 0:
                return w / OXIDE_MW[oxide] * OXIDE_CAT[oxide] * scale
        except (ValueError, TypeError):
            pass
    return np.nan


# ── Sample name mapping ──────────────────────────────────────────────────────
# Maps each entry in sample_names.txt to the prefix(es) used in data files.
SAMPLE_PREFIXES = {
    '19NC08BD': ['19NC08B'],
    '19NC06BC': ['19NC6D', '19NC06D'],   # BC file uses 19NC6D; Amp/Pl XLS uses 19NC06D
    '19NC11AC': ['19NC11C'],
    'YG182B':   ['YG182'],
}


def get_prefix(sample_line):
    """Extract the sample code (first token before space/dash)."""
    code = sample_line.split('–')[0].split('-')[0].strip()
    return code.split()[0] if code else ''


def name_matches(data_name, sample_code):
    prefixes = SAMPLE_PREFIXES.get(sample_code, [sample_code])
    return any(data_name.startswith(p) or p in data_name for p in prefixes)


def safe_float(v):
    try:
        return float(v)
    except (ValueError, TypeError):
        return np.nan


# ── File discovery ───────────────────────────────────────────────────────────
all_files = os.listdir('.')
xlsx_files = [f for f in all_files if f.endswith('.xlsx')]
xls_files  = [f for f in all_files if f.endswith('.xls') and not f.endswith('.xlsx')]

bc_file  = next((f for f in xlsx_files if 'enki' in f.lower()), None)
amp_xls  = next((f for f in xls_files if 'Amphibole-norm1-serie1' in f and 'Toulouse' not in f), None)
amp_T    = next((f for f in xlsx_files if 'Amphibole' in f and 'Toulouse' in f), None)
pl_xls   = next((f for f in xls_files if 'Gen-Oxide-norm-Pl' in f), None)
cpx_xls  = next((f for f in xls_files if 'Pyroxene-norm' in f and 'Toulouse' not in f), None)
grt_xlsx = next((f for f in xlsx_files if 'GRT' in f), None)

print(f'Using files:\n  Pl:  {pl_xls}\n  Amp: {amp_xls}\n  AmpT:{amp_T}\n  Cpx: {cpx_xls}\n  Ep:  {bc_file}\n  Grt: {grt_xlsx}')


# ── Read Pl_Ca ───────────────────────────────────────────────────────────────
def read_pl_ca():
    """Returns [(sample_name, Ca_apfu), ...] from General oxygen sheet, Atoms/Ca row."""
    wb = xlrd.open_workbook(pl_xls)
    sh = wb.sheet_by_name('General oxygen')
    atoms_row = next((i for i in range(sh.nrows) if str(sh.cell_value(i, 0)).strip() == 'Atoms'), None)
    ca_row    = next((i for i in range(atoms_row + 1, min(atoms_row + 20, sh.nrows))
                      if str(sh.cell_value(i, 0)).strip() == 'Ca'), None)
    input_row = next((i for i in range(sh.nrows) if str(sh.cell_value(i, 0)).strip() == 'Input'), None)
    result = []
    for j in range(1, sh.ncols):
        n = str(sh.cell_value(input_row, j)).strip()
        if n and n != '0.0':
            v = safe_float(sh.cell_value(ca_row, j))
            if not np.isnan(v):
                result.append((n, v))
    return result


# ── Read Amp_Mg ──────────────────────────────────────────────────────────────
def read_amp_mg():
    """Returns [(sample_name, Mg_apfu), ...] from NAMP sheet, Cations all Fe2+/Mg row."""
    result = []

    def _read_xls(wb, shname):
        sh = wb.sheet_by_name(shname)
        input_row = next((i for i in range(sh.nrows) if str(sh.cell_value(i, 0)).strip() == 'Input'), None)
        cats_row  = next((i for i in range(sh.nrows) if 'Cations all Fe2+' in str(sh.cell_value(i, 0))), None)
        mg_row    = next((i for i in range(cats_row + 1, min(cats_row + 20, sh.nrows))
                          if str(sh.cell_value(i, 0)).strip() == 'Mg'), None)
        for j in range(1, sh.ncols):
            n = str(sh.cell_value(input_row, j)).strip()
            if n and n not in ('0.0', 'Average'):
                v = safe_float(sh.cell_value(mg_row, j))
                if not np.isnan(v):
                    result.append((n, v))

    _read_xls(xlrd.open_workbook(amp_xls), 'NAMP')

    wb_T = openpyxl.load_workbook(amp_T, read_only=True, data_only=True)
    T_rows = list(wb_T['NAMP'].iter_rows(values_only=True))
    wb_T.close()
    cats_row_T = next((i for i, row in enumerate(T_rows) if row and 'Cations all Fe2+' in str(row[0])), None)
    mg_row_T   = next((i for i in range(cats_row_T + 1, min(cats_row_T + 20, len(T_rows)))
                       if T_rows[i] and str(T_rows[i][0]).strip() == 'Mg'), None)
    inp_row_T  = next((i for i, row in enumerate(T_rows) if row and str(row[0]).strip() == 'Input'), None)
    name_row_T = inp_row_T + 1 if inp_row_T is not None else None

    if name_row_T is not None and mg_row_T is not None:
        names = T_rows[name_row_T]
        mgs   = T_rows[mg_row_T]
        for j in range(1, len(names)):
            n = str(names[j]).strip() if names[j] is not None else ''
            if n and n not in ('None', '0.0', 'Average'):
                v = safe_float(mgs[j] if j < len(mgs) else None)
                if not np.isnan(v):
                    result.append((n, v))

    return result


# ── Read Amp_Si ──────────────────────────────────────────────────────────────
def read_amp_si():
    """Returns [(sample_name, Si_apfu), ...] from NAMP sheet, Cations all Fe2+/Si row.
    Primary sources only — no BC supplement."""
    result = []

    def _read_xls(wb, shname):
        sh = wb.sheet_by_name(shname)
        input_row = next((i for i in range(sh.nrows) if str(sh.cell_value(i, 0)).strip() == 'Input'), None)
        cats_row  = next((i for i in range(sh.nrows) if 'Cations all Fe2+' in str(sh.cell_value(i, 0))), None)
        si_row    = next((i for i in range(cats_row + 1, min(cats_row + 20, sh.nrows))
                          if str(sh.cell_value(i, 0)).strip() == 'Si'), None)
        for j in range(1, sh.ncols):
            n = str(sh.cell_value(input_row, j)).strip()
            if n and n not in ('0.0', 'Average'):
                v = safe_float(sh.cell_value(si_row, j))
                if not np.isnan(v):
                    result.append((n, v))

    _read_xls(xlrd.open_workbook(amp_xls), 'NAMP')

    wb_T = openpyxl.load_workbook(amp_T, read_only=True, data_only=True)
    T_rows = list(wb_T['NAMP'].iter_rows(values_only=True))
    wb_T.close()
    cats_row_T = next((i for i, row in enumerate(T_rows) if row and 'Cations all Fe2+' in str(row[0])), None)
    si_row_T   = next((i for i in range(cats_row_T + 1, min(cats_row_T + 20, len(T_rows)))
                       if T_rows[i] and str(T_rows[i][0]).strip() == 'Si'), None)
    inp_row_T  = next((i for i, row in enumerate(T_rows) if row and str(row[0]).strip() == 'Input'), None)
    name_row_T = inp_row_T + 1 if inp_row_T is not None else None

    if name_row_T is not None and si_row_T is not None:
        names = T_rows[name_row_T]
        sis   = T_rows[si_row_T]
        for j in range(1, len(names)):
            n = str(names[j]).strip() if names[j] is not None else ''
            if n and n not in ('None', '0.0', 'Average'):
                v = safe_float(sis[j] if j < len(sis) else None)
                if not np.isnan(v):
                    result.append((n, v))

    return result


# ── Read Aug_Mg ──────────────────────────────────────────────────────────────
def read_cpx_mg():
    """Returns [(sample_name, Mg_apfu), ...] from WB-cpx sheet, Cations all Fe2+/Mg row."""
    wb = xlrd.open_workbook(cpx_xls)
    sh = wb.sheet_by_name('WB-cpx')
    input_row = next((i for i in range(sh.nrows) if str(sh.cell_value(i, 0)).strip() == 'Input'), None)
    cats_row  = next((i for i in range(sh.nrows) if 'Cations all Fe2+' in str(sh.cell_value(i, 0))), None)
    mg_row    = next((i for i in range(cats_row + 1, min(cats_row + 20, sh.nrows))
                      if str(sh.cell_value(i, 0)).strip() == 'Mg'), None)
    result = []
    for j in range(1, sh.ncols):
        n = str(sh.cell_value(input_row, j)).strip()
        if not n or n == '0.0' or re.match(r'^[\d.eE\-+]+$', n) or '/ 1' in n:
            continue
        v = safe_float(sh.cell_value(mg_row, j))
        if not np.isnan(v) and v != 0.0:
            result.append((n, v))
    return result


# ── Read Ep oxide data ────────────────────────────────────────────────────────
def read_ep_ca():
    """Returns [(sample_name, Ca_apfu), ...] calculated from Feuil2 EP section (12.5 O).
    Preserves duplicate analysis names as separate entries."""
    wb = openpyxl.load_workbook(bc_file, read_only=True, data_only=True)
    rows = list(wb['Feuil2'].iter_rows(values_only=True))
    wb.close()

    ep_start = next((i for i, r in enumerate(rows) if r and str(r[0]).strip() == 'EP'), None)
    comment_row_idx = ep_start + 2

    comments = rows[comment_row_idx]
    # Build list of (col_index, name) preserving duplicates
    ep_cols = [(j, str(comments[j]).strip())
               for j in range(1, len(comments))
               if comments[j] and '_Ep' in str(comments[j])]

    ep_oxides = {j: {} for j, _ in ep_cols}
    for ri in range(comment_row_idx + 1, comment_row_idx + 20):
        if ri >= len(rows):
            break
        row = rows[ri]
        ox = str(row[0]).strip() if row[0] is not None else ''
        if ox == 'F ':
            ox = 'F'
        if ox in OXIDE_MW:
            for j, _ in ep_cols:
                v = row[j] if j < len(row) else None
                try:
                    ep_oxides[j][ox] = float(v) if v is not None else 0.0
                except (ValueError, TypeError):
                    ep_oxides[j][ox] = 0.0

    return [(name, calc_apfu_from_oxides(ep_oxides[j], 12.5, 'Ca'))
            for j, name in ep_cols]


# ── Read any Feuil2 mineral section from oxides ──────────────────────────────
def _read_feuil2_section(rows, section_label):
    """Return list of (col_idx, sample_name, {oxide: wt%}) for a named section."""
    STOP_LABELS = {'AMPH', 'EP', 'PREHNITE', 'PX', 'PL', 'GRT', '????'}
    sec_row = next((i for i, r in enumerate(rows)
                    if r and str(r[0]).strip() == section_label
                    and not any(v for v in r[1:4] if v)), None)
    if sec_row is None:
        return []
    comment_row_idx = sec_row + 2
    comments = rows[comment_row_idx]
    cols = [(j, str(comments[j]).strip())
            for j in range(1, len(comments))
            if comments[j] and str(comments[j]).strip() not in ('None', '')]
    data = {j: {} for j, _ in cols}
    for ri in range(comment_row_idx + 1, len(rows)):
        row = rows[ri]
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        if label in STOP_LABELS and not any(v for v in row[1:4] if v):
            break
        if label == 'F ':
            label = 'F'
        if label in OXIDE_MW:
            for j, _ in cols:
                v = row[j] if j < len(row) else None
                try:
                    data[j][label] = float(v) if v is not None else 0.0
                except (ValueError, TypeError):
                    data[j][label] = 0.0
        if label == 'Total':
            break
    return [(j, name, data[j]) for j, name in cols]


def read_amp_mg_from_bc():
    """Calculate Amp_Mg (23 O) from Feuil2 AMPH section oxide data.
    Returns [(sample_name, Mg_apfu), ...]."""
    wb = openpyxl.load_workbook(bc_file, read_only=True, data_only=True)
    rows = list(wb['Feuil2'].iter_rows(values_only=True))
    wb.close()
    entries = _read_feuil2_section(rows, 'AMPH')
    return [(name, calc_apfu_from_oxides(oxides, 23.0, 'Mg'))
            for _, name, oxides in entries
            if '_Amph' in name or '_Amp' in name]


def read_amp_si_from_bc():
    """Calculate Amp_Si (23 O) from Feuil2 AMPH section oxide data.
    Returns [(sample_name, Si_apfu), ...]."""
    wb = openpyxl.load_workbook(bc_file, read_only=True, data_only=True)
    rows = list(wb['Feuil2'].iter_rows(values_only=True))
    wb.close()
    entries = _read_feuil2_section(rows, 'AMPH')
    return [(name, calc_apfu_from_oxides(oxides, 23.0, 'Si'))
            for _, name, oxides in entries
            if '_Amph' in name or '_Amp' in name]


def read_pl_ca_from_bc():
    """Calculate Pl_Ca (8 O) from Feuil2 PL section oxide data.
    Returns [(sample_name, Ca_apfu), ...]."""
    wb = openpyxl.load_workbook(bc_file, read_only=True, data_only=True)
    rows = list(wb['Feuil2'].iter_rows(values_only=True))
    wb.close()
    entries = _read_feuil2_section(rows, 'PL')
    return [(name, calc_apfu_from_oxides(oxides, 8.0, 'Ca'))
            for _, name, oxides in entries
            if '_Pl' in name]


# ── Read Grt_Mg ──────────────────────────────────────────────────────────────
def read_grt_mg():
    """Returns [(sample_name, Mg_apfu), ...] from Gen-Oxide-norm-GRT.xlsx,
    General oxygen sheet, Atoms/Mg row. Sample names are in the row after Input."""
    if grt_xlsx is None:
        return []
    wb = openpyxl.load_workbook(grt_xlsx, read_only=True, data_only=True)
    rows = list(wb['General oxygen'].iter_rows(values_only=True))
    wb.close()
    input_row = next((i for i, r in enumerate(rows) if r and str(r[0]).strip() == 'Input'), None)
    name_row  = input_row + 1 if input_row is not None else None
    atoms_row = next((i for i, r in enumerate(rows) if r and str(r[0]).strip() == 'Atoms'), None)
    mg_row    = next((i for i in range(atoms_row + 1, min(atoms_row + 20, len(rows)))
                      if rows[i] and str(rows[i][0]).strip() == 'Mg'), None)
    if name_row is None or mg_row is None:
        return []
    names = rows[name_row]
    mgs   = rows[mg_row]
    result = []
    for j in range(1, len(names)):
        n = str(names[j]).strip() if names[j] is not None else ''
        if not n or n in ('None', '0.0'):
            continue
        v = safe_float(mgs[j] if j < len(mgs) else None)
        if not np.isnan(v):
            result.append((n, v))
    return result


# ── Aggregate by sample ───────────────────────────────────────────────────────
def aggregate(data_pairs, sample_code):
    """Collect all values matching sample_code prefix, return (mean, std).
    data_pairs: list of (name, value) tuples
    """
    vals = [v for name, v in data_pairs if name_matches(name, sample_code)]
    if not vals:
        return np.nan, np.nan
    arr = np.array(vals, dtype=float)
    return float(np.nanmean(arr)), float(np.nanstd(arr, ddof=0))


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    with open('sample_names.txt', 'r', encoding='utf-8') as fh:
        sample_lines = [ln.strip() for ln in fh if ln.strip()]

    print('Reading data sources...')
    pl_data   = read_pl_ca()
    amp_data  = read_amp_mg()
    amp_si_data = read_amp_si()
    cpx_data  = read_cpx_mg()
    ep_data   = read_ep_ca()
    grt_data  = read_grt_mg()

    # Supplement with oxide-calculated values from BenedicteCenki Feuil2.
    # Only add BC data for samples that have no primary measurements,
    # to avoid mixing datasets.
    def covered_by(data_list, sample_code):
        return any(name_matches(n, sample_code) for n, _ in data_list)

    bc_amp    = read_amp_mg_from_bc()
    bc_amp_si = read_amp_si_from_bc()
    bc_pl     = read_pl_ca_from_bc()

    for line in sample_lines:
        code = get_prefix(line)
        if not covered_by(amp_data, code):
            amp_data    += [(n, v) for n, v in bc_amp    if name_matches(n, code)]
        if not covered_by(amp_si_data, code):
            amp_si_data += [(n, v) for n, v in bc_amp_si if name_matches(n, code)]
        if not covered_by(pl_data, code):
            pl_data     += [(n, v) for n, v in bc_pl     if name_matches(n, code)]

    print(f'  Pl_Ca  : {len(pl_data)} measurements')
    print(f'  Amp_Mg : {len(amp_data)} measurements')
    print(f'  Amp_Si : {len(amp_si_data)} measurements')
    print(f'  Aug_Mg : {len(cpx_data)} measurements')
    print(f'  Ep_Ca  : {len(ep_data)} measurements')
    print(f'  Grt_Mg : {len(grt_data)} measurements')

    ALL_COLS = ['Pl_Ca', 'Amp_Mg', 'Amp_Si', 'Aug_Mg', 'Ep_Ca', 'Grt_Mg']
    ALL_DATA = [pl_data, amp_data, amp_si_data, cpx_data, ep_data, grt_data]

    for line in sample_lines:
        sample_code = get_prefix(line)
        safe_name = re.sub(r'[^\w\-]', '_', line).strip('_')

        means, stds, cols = [], [], []
        for col, data in zip(ALL_COLS, ALL_DATA):
            m, s = aggregate(data, sample_code)
            if not np.isnan(m):
                cols.append(col)
                means.append(m)
                stds.append(s)

        def fmt(v):
            return '' if np.isnan(v) else f'{v:.6f}'

        header   = ','.join(cols)
        mean_row = ','.join(fmt(v) for v in means)
        std_row  = ','.join(fmt(v) for v in stds)

        outfile = f'{safe_name}.csv'
        with open(outfile, 'w') as fo:
            fo.write(header + '\n')
            fo.write(mean_row + '\n')
            fo.write(std_row + '\n')

        print(f'\n{line}  →  {outfile}')
        print(f'  cols: {header}')
        print(f'  mean: {mean_row}')
        print(f'  std:  {std_row}')


if __name__ == '__main__':
    main()
